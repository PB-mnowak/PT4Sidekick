# TODO Get protein data if ID given
# Log: adding stocks - protein / stock / location
# Tag detection

# TODO Sort helper functions

from string import ascii_lowercase, ascii_uppercase
from os import listdir, getcwd, system
from os.path import isfile, join
from getpass import getpass
import warnings

from Bio import GenBank, BiopythonWarning
from Bio.Seq import Seq
from openpyxl import load_workbook, Workbook
import pandas as pd
import re
import requests

from src.pt4_protein import Protein
from src.pt4_data import *

# Main menu

def main():
    system('cls')
    global token
    global pb_all
    global test_mode 
    
    mode_selection = input('Test mode (Y/N): ').lower()
    test_mode = mode_selection == 'y'
    token = get_token(test_mode=test_mode)
    
    pb_all = 'P:\\_research group folders\\PT Proteins\\PT4\\_PT4_sidekick'

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    while True:
        system('cls')
        menu_dict = {
            '1': ['Create HTsyn template file', 
                generate_template],
            '2': ['Collect stock data from Labguru',
                get_plasmid_data],
            '3': ['Calculate protein parameters from GeneBank file',
                protein_analysis],
            '4': ['Create Labguru import file for plasmid stock transfer',
                create_import_file],
            '5': ['Add new proteins and their stocks to Labguru',
                add_pt_stocks],
            '6': ['Create excel file for label printing',
                create_label_xlsx]
            }

        # Print menu
        print_menu(menu_dict)

        user_input = input('\nTask number: ')
        
        if user_input.lower() == "q":
            break
        elif user_input in menu_dict:
            system('cls')
            menu_dict[user_input][1]()
            system('pause')
        else:
            print('---< Wrong input. Try again >---'.center(80))
            system('pause')

# def timeis(func):
#     """Decorator that reports the execution time."""
#     def wrap(*args, **kwargs):
#         start = time.time()
#         result = func(*args, **kwargs)
#         end = time.time()
        
#         print(f'Executed in {end-start:.3f} s')
#         return result
#     return wrap

# Helpers

def print_menu(menu_dict):
        print('/ PT4 Sidekick \\'.center(80, '_'), '\n')
        for k, v in menu_dict.items():
            print(f'{k:>2}) {v[0]}')
        print('\n Q - Quit')
        print(''.center(80, '_'))

def task_start(file = None):
    print("/ Task in progress \\".center(80, '_'), '\n', sep='')
    if file is not None:
        print(f'File: {file}\n')

def task_end():
    print("\n", "\\ Task completed /".center(80, '‾'), '\n', sep="")

def pball_connection(file):
    while True:
        if isfile(join(pb_all, file)):
            break        
        else:
            print('---< Template file not found: check connection to PB_all >---'.center(80))
            system('pause')

def verify_file_sheets(path: str, file: str):
    """ Verify if main sheets are present in the file """
    
    wb = load_workbook(join(path, file))
    
    try:
        wb_sheets = set(wb.sheetnames)
        pl_columns_set = set([cell.value for cell in wb['Plasmids'][1]])
        pt_columns_set = set([cell.value for cell in wb['Proteins'][1]])
        
        wb.close()
       
        all_sheets = wb_sheets.issuperset(sheet_verification['sheets'])
        pl_all_cols, pt_all_cols = False, False
        
        if all_sheets:
            pl_columns_set = set([cell.value for cell in wb['Plasmids'][1]])
            pt_columns_set = set([cell.value for cell in wb['Proteins'][1]])
            
            pl_all_cols = pl_columns_set.issuperset(sheet_verification['Plasmids'])
            pt_all_cols = pt_columns_set.issuperset(sheet_verification['Proteins'])
        
        conditions =  [all_sheets, pl_all_cols, pt_all_cols]
        
        if all(conditions):
            return True
        else:
            issues = [
                'Missing workbook sheets.',
                'Missing columns in "Plasmids" sheet.',
                'Missing columns in "Proteins" sheet.'
                ]
            
            detected_issues = [issue for issue, cond in zip(issues, conditions) if not cond]
            print(f'---< {" ".join(detected_issues)} Select other file >---'.center(80))
            return False
         
    except Exception as e:
        print(e)
        wb.close()
        return False
    

def print_task(func):
    def wrap(*args, **kwargs):
        task_start()
        result = func(*args, **kwargs)
        task_end()
        return result
    return wrap

def if_overwrite():
    while True:
        overwrite = input("Overwrite existing records? (Y/N) ").upper()
        if overwrite == "Y":
            print()
            return True
        elif overwrite == "N":
            return False
        else:
            print("---< Wrong input. Try again >---".center(80))

def get_path_file(ext):
    mypath = getcwd()
    xlsx_list = scan_files(mypath, ext)
    file = choose_file(mypath, xlsx_list, ext)
    if file is None:
        return mypath, None
    print()
    return mypath, file

def scan_files(path, ext):
    """ Returns list of xlsx files in given path """
    xlsx_list = [file for file in listdir(path) if check_filename(path, file, ext)]
    return xlsx_list

def check_filename(path, file, ext):
    """  """
    is_file = isfile(join(path, file))
    is_xlsx = file.endswith(f'.{ext}')
    not_temp = not file.startswith('~$')
    conditions = [is_file, is_xlsx, not_temp]
    return all(conditions)

def choose_file(path, file_list, ext):
    print('/ Select file \\'.center(80, '_'), '\n', sep='')
    xlsx_dict = {str(i): file for i, file in enumerate(file_list, 1)}    
    for i, file in xlsx_dict.items():
        print(f'{i.rjust(2)})', file.replace(f'.{ext}', ''), sep=" ")
    print('\n', 'Q - Return to menu')
    print(''.center(80, '_'), '\n', sep='')
    
    while True:
        file_i = input('File number: ').lower()
        if file_i in xlsx_dict:
            if verify_file_sheets(path, xlsx_dict[file_i]):
                break
        elif file_i == 'q':
            return None
        else:
            print('---< Wrong input. Try again >---'.center(80))

    system('cls')

    return xlsx_dict[file_i]

def scan_genebank(path):
    gb_files = [f for f in listdir(path) if isfile(join(path, f)) and f.rpartition(".")[-1] == "gb"]
    return gb_files

def pos_to_int(pos_str):
    """ Changes position naming between index and LetterNumber formats """
    
    def is_alnum(pos_str):
        try:
            pos_split = list(pos_str)
            cond_1 = pos_split == 2
            cond_2 = pos_split[0].isalpha()
            cond_3 = pos_split[0].isnumeric()
            conditions = [cond_1, cond_2, cond_3]
            return all(conditions)
        except:
            return False
    
    # TODO pos <= 81
    
    try:
        if is_alnum(pos_str):
            pos_split = list(pos_str)
            row = ord(pos_split[0].upper()) - 65
            col = int(pos_split[1])
            return row * 9 + col
        else:
            pos_int = int(pos_str.split(' ')[0])
        return pos_int
    except:
        print(f'---< Wrong type of stock position format: {pos_str} >---'.center(80))
        print("Accepted formats: '1 (A1)', '1' or 'A1'")

def pos_to_str(pos_int):
    try:
        row, col = divmod(pos_int-1, 9)
        row_alpha = chr(65 + row)
        return f'{pos_int} ({row_alpha}{col+1})'
    except:
        print(f'---< Wrong type of stock position format: {type(pos_int)} >---'.center(80))

def sheet_header(row):
    # for c in row:
    #     print(c.value, end=" | ")
    header = [(cell.value, i) for i, cell in enumerate(row, 1) if cell.value is not None]
    
    return dict(header)

def save_workbook(wb, mypath, file):
    while True:
        try:
            wb.save(join(mypath, file))
            print(f'\nFile updated: {file}')
            break
        except PermissionError:
            print(f'---< Unable to save changes in {file} >---'.center(80))
            print('Close the file and continue'.center(80))
            system('pause')

def write_data(sheet, header_dict, data, i):
    '''
    Writes data to cell in the sheet:
        sheet:          workbook.sheet
        header_dict:    column_name: column_index
        data:           value, column_name
        i:              row index
    '''
    for value, column_name in data:
        try:
            sheet.cell(i, header_dict[column_name]).value = value
        except Exception as e:
            print(e)

def write_hyperlinks(sheet, header_dict, urls, i):
    '''
    Writes data to cell in the sheet:
        sheet:          workbook.sheet
        header_dict:    column_name: column_index
        data:           value, column_name
        i:              row index
    '''
    for url, column_name in urls:
        try:
            sheet.cell(i, header_dict[column_name]).hyperlink = url
        except Exception as e:
            print(e)

def choose_box_name(box_name):
    print(f'\nSuggested new name: {box_name}')
    while True:
        box_name_approve = input('Y - accept | N - change').upper()
        if box_name_approve == 'Y':
            break
        elif  box_name_approve == 'N':
            box_name = input('Enter box name: ')
        else:
            print('---< Wrong input. Try again >---'.center(80))
    return box_name

def conc_conversion(concentration, unit):
    try:
        mass_unit, vol_unit = unit.strip().split('/')
        mass = conc_units.get(mass_unit, 1)
        vol = conc_units.get(vol_unit, 1)
        if concentration is not None:
            concentration = float(concentration) * int(mass / vol)
    except:
        pass
    return concentration

def volume_conversion(volume, lg_unit, base_unit='μl'):
    try:
        unit_value = lg_units_conversion[lg_unit]
        base_unit_value = conc_units[base_unit]
        if volume is not None:
            volume = float(volume) * int(unit_value / base_unit_value)
    except KeyError:
        pass
    return volume

def gen_prot_name(plasmid_inv_name: str) -> str:
    """
    Generates protein inventory from plasmid inventory name by
    removing plasmid type
    """
    try:
        pattern = f'[{ascii_lowercase}]+[{ascii_uppercase}]+[._]'
        re_obj = re.compile(pattern)
        return re_obj.sub('', plasmid_inv_name)
    except Exception as e:
        print(e)
        return plasmid_inv_name


# API requests

def get_token(token = None, test_mode=False):
    if token is None:
        print("\nEnter Labguru credentials: name (n.surname) and password")
        while True:

            if test_mode is True:
                name = ''
                password = ''
            else:
                name = str(input("Name: ")).lower() + "@purebiologics.com"
                password = str(getpass("Password: "))
            
            body = {"login": name, "password": password}
            
            message = "\nAcquiring authentication token"
            print(f"{message.ljust(73, '.')}", end="")
            
            session = requests.post("https://my.labguru.com/api/v1/sessions.json", body)
            token = session.json()['token']

            if token and token != "-1":
                print("SUCCESS")
                break
            else:
                print("..ERROR")
                print("Wrong login or password. Try again\n")
    return token

def get_hierarchy(token, storage_id):
    url = f'https://my.labguru.com/api/v1/storages/{storage_id}.json?token={token}'
    session = requests.get(url)
    storage = session.json()
    hierarchy = storage['name_with_hierarchy']
    return hierarchy

def get_sysid_pl(pl_id: str):
    """ Gets SysID of plasmid based on its ID """
    url = f'https://my.labguru.com/api/v1/plasmids/{pl_id}.json?token={token}'
    session = requests.get(url)
    return session.json()['sys_id']

def scan_storage(token: str, storage_id: str) -> dict:
    """ 
    Returns two dicts of boxes ID and their names based on Labguru JSON record
    of the given storage ID:
    1) boxes_sti (string to ID):    {name: id}  
    2) boxes_its (ID to string):    {id: name} 

    Box Z23.C.01 (ID: 1004) is excluded - only for positive controls
    """
    url = f'https://my.labguru.com/storage/storages/{storage_id}?token={token}'
    page = requests.get(url)
    
    storage = page.json()
    boxes_sti = {box['name']: box['id'] for box in storage['boxes'] if box['id'] != 1004}
    boxes_its = {box['id']: box['name'] for box in storage['boxes'] if box['id'] != 1004}

    return boxes_sti, boxes_its

def get_box_data(token: str, box_id: str): # token
    '''
    Returns dict of data from given box record
    '''
    url = f'https://my.labguru.com/api/v1/boxes/{box_id}.json?token={token}'
    session = requests.get(url)
    box = session.json()

    box_occupied_pos = []
    size = int(box["rows"]) * int(box["cols"])
    for stock in box['stocks']:
        box_occupied_pos.append(stock['stock']['location_in_box'])

    box_data = {
        'name': box['name'],
        'url': f'/storage/boxes/{box_id}',
        'title': box["name"],
        'rows': box["rows"],
        'cols': box["cols"],
        'storage_id': str(box["storage_id"]),
        'hierarchy': get_hierarchy(token, str(box['storage_id'])),
        'free_pos': [i for i in range(1, size+1) if i not in box_occupied_pos]
    }
    
    return box_data

# TODO Create new box if no box found
def add_box(box_types=box_types, storage=None):
    
    if storage is None:
        boxes_str, boxes_id = scan_storage(token, '1796')
        storage = boxes_str
       
    box_num = max([int(box[-2:]) for box in storage.keys()]) + 1
    box_name = f'Z23.C.{box_num:02d}'
    box_rows, box_cols = 9, 9
    
    box_name = choose_box_name(box_name)
    
    url = 'https://my.labguru.com/api/v1/boxes'
    body = {'token': token,
            'item': {
                "name": box_name,
                "rows": box_rows,
                "cols": box_cols,
                "box_type": box_types['Plasmid'],
                "color": 'yellow',
                "shared": "1"
                }
            }

    session = requests.post(url, json=body)
    
    if session.status_code == 201:
        new_box = session.json()
        # TODO - open box page to change the box location to 4.38
        system(f'explorer "https://my.labguru.com/storage/boxes/{new_box["id"]}/edit"')
        return new_box
    else:
            print(f'Error while handling {box_name} - Code {session.status_code}')

def add_protein(prot_data):
    """ """
    url = 'https://my.labguru.com/api/v1/proteins'
    body = {"token": token,
            "item": prot_data}

    session = requests.post(url, json=body)
    
    # for d in dir(session):
    #     print(d, eval(f'session.{d}'), sep='\t')
        
    if session.status_code == 201:
        prot = session.json()
        return prot
    else:
        print(f'Error while handling {prot_data["name"]} - Code {session.status_code}')

def add_stock(stock_data):
    """ """
    url = 'https://my.labguru.com/api/v1/stocks'
    body = {'token': token,
            'item': stock_data}

    session = requests.post(url, json=body)
    if session.status_code == 201:
        stock = session.json()
        return stock
    else:
            print(f'Error while handling {stock_data["name"]} - Code {session.status_code}')


# 1) New template

@print_task
def generate_template():
    """ Create copy of template file stored on PB_all """
    template_file = "Templates\\HTsyn_template.xlsx"
    pball_connection(template_file)
    
    name = input('Name of the template file: ') + ".xlsx"
    mypath = getcwd()
    while name in listdir(mypath):
        print('---< A duplicate file name exists >---'.center(80))
        name = input('Select other name: ') + ".xlsx"
        
    system(f'copy "{join(pb_all, template_file)}" "./{name}" >nul')
    print(f'\nTemplate file {name} created')  


# 2) Load plasmid data

def get_plasmid_data():

    def collect_ids(overwrite=False):
        stock_ids = []
        for i in range(2,102):
            stock_id = ws_plasmids.cell(
                column=pl_header['ID'],
                row=i
                ).value
            if not overwrite:
                plasmid_name = ws_plasmids.cell(
                column=pl_header['Plasmid inventory name'],
                row=i
                ).value
                if plasmid_name is not None:
                    continue
            if stock_id is not None:
                stock_ids.append((i, str(stock_id), ))
        return stock_ids

    mypath, file = get_path_file('xlsx')
    if file is None:
        return None

    task_start(file)
    
    overwrite = if_overwrite()
    
    lg_url = 'https://my.labguru.com/'
    filler = '.'

    wb = load_workbook(join(mypath, file))

    ws_plasmids = wb["Plasmids"]
    ws_proteins = wb["Proteins"]

    pl_header = sheet_header(ws_plasmids[1])
    pt_header = sheet_header(ws_proteins[1])

    # with open('headers.txt', 'w+') as file:

    #     for col, i in pl_header.items():
    #         file.write(f'{col.ljust(30)}{i}\n')
    #     for col, i in pt_header.items():
    #         file.write(f'{col.ljust(30)}{i}\n')

    stock_ids = collect_ids(overwrite)

    print("Downloading stock data:\n")

    for i, stock in stock_ids:
        print(f"{i-1:>3}) Stock {stock:{filler}<63}", end="")
        stock_url = f'https://my.labguru.com/api/v1/stocks/{stock}.json?token={token}'
        record = requests.get(stock_url)
        record_json = record.json()
        
        # TODO >>> Function
        try:
            if record_json["content_type_for_display"] != 'Plasmid':
                print(f"FAILED")
                print(f'---< Stock {stock} is not a plasmid >---'.center(80))
                continue
        except KeyError:
            print(f"FAILED")
            continue

        plasmid_name = record_json["name"]
        plasmid_id = record_json["stockable"]["id"]
        plasmid_sysid = get_sysid_pl(plasmid_id)
        plasmid_inv_name = record_json["stockable"]["name"]
        concentration = record_json["concentration"]
        if concentration:
            concentration = float(concentration)
        volume_unit = record_json['volume_unit_id']
        volume = record_json['volume']
        volume = volume_conversion(volume, volume_unit, 'μl')
        description = record_json["description"]
        try:
            box = record_json["storage"]["name"]
            position = record_json["box"]["location_in_box"]
        except TypeError:
            box = 'Consumed'
            position = ''
        stock_url = lg_url + record_json["url"]
        inv_url = lg_url + record_json["stockable"]["url"]
        link = 'LINK'
        
        if not box.startswith('Z23.C.'):
            transfer = 'Y'
        else:
            transfer = ''
        
        prot_name = gen_prot_name(plasmid_inv_name)

        prot_i = 2 + (i-2) * 5
        
        if description:
            try:
                conc_pattern = r'[cC]\w*.? *= *[0-9|.]+ *.g\/.[lL]'
                re_obj = re.compile(conc_pattern)
                match = re_obj.search(description)
                if match is not None:
                    conc_raw = match.group()
                    conc_unit = conc_raw.partition('=')[2]
                    concentration = float(conc_unit[:-5].strip())
                    unit = conc_unit[-5:].lower().replace('u', 'μ')
                    concentration = conc_conversion(concentration, unit)
            except Exception as e:
                print(e)
        
        if 'filt' in description.lower():
            filtration = 'Y'
        else:
            filtration = 'N'
        
        plasmid_data = [
            (plasmid_name, 'Stock name (Plasmid)'),
            (filtration, 'Filtered'),
            (concentration, 'Conc'),
            (volume, 'Volume'),
            (box, 'Box'),
            (position, 'Position'),
            (link, 'Link - Stock'),
            (plasmid_sysid, 'SysID'),
            (plasmid_inv_name, 'Plasmid inventory name'),
            (link, 'Link - Inv'),
            (transfer, 'Transfer'),
        ]
        
        plasmid_hyperlinks = [
            (stock_url, 'Link - Stock'),
            (inv_url, 'Link - Inv'), 
        ]
        
        prot_data = [
            (prot_name, 'POI name'),
            (plasmid_sysid, 'Plasmid SysID'),
            (plasmid_inv_name, 'Plasmid inventory name'),
        ]
        
        # TODO - consumed stock (None)
        
        # Plasmids sheet
        write_data(ws_plasmids, pl_header, plasmid_data, i)
        write_data(ws_proteins, pt_header, prot_data, prot_i)
        write_hyperlinks(ws_plasmids, pl_header, plasmid_hyperlinks, i)

        if record.status_code == 200:
            print(f"..DONE")
        else:
            print(f"FAILED")
    
    save_workbook(wb, mypath, file)
    wb.close()
    
    task_end()


# 3) Protein analysis

def protein_analysis():
    """ """
    
    warnings.filterwarnings('ignore', category=BiopythonWarning, module='Bio')
    
    mypath, file = get_path_file('xlsx')
    
    if file is None:
        return None

    ss_list = protein_db['ss']
    end_list = protein_db['end']
    
    # TODO Scan db for matches
    # TODO Select if 1x or 2x

    ss_1 = "ATGTACAGGATGCAACTCCTGTCTTGCATTGCACTAAGTCTTGCACTTGTCACGAATTCA"
    ss_2 = "ATGTACAGGATGCAACTCCTGTCTTGCATTGCACTAAGTCTTGCACTTGTCACGAATTCG"
    light_end = "ACAAAGAGCTTCAACAGGGGAGAGTGTTAG"
    len_light_end = len(light_end)
    heavy_end = "AAGAGCCTCTCCCTGTCTCCGGGTAAATGA" # IgG1
    # heavy_end = "GCCTCTCCCTGTCTCCGGGTAAATAAAC" # IgG4
    len_heavy_end = len(heavy_end)
    filler = '.'
    
    task_start(file)

    wb = load_workbook(join(mypath, file))
    ws_proteins = wb["Proteins"]
    pt_header = sheet_header(ws_proteins[1])
    gb_files = scan_genebank(mypath)

    # No gb files detected
    if not gb_files:
        print("\nNo GeneBank files found - protein analysis omitted\n")
        task_end()
        pass

    print(f"{len(gb_files)} GeneBank files detected\n")
    print(f"Evaluating protein parameters{'':{filler}<47}", end="")

    found_seqs = 0
    
    # TODO
    # refactor - gb list, protein_list
    
    for gb_file in gb_files:
        with open(join(mypath, gb_file)) as handle:          
            for record in GenBank.parse(handle):
                for i in range(2, 502, 5):
                    if (plasmid_inv_name := ws_proteins.cell(
                            column=pt_header['Plasmid inventory name'],
                            row=i
                            ).value) is None:
                        continue

                    elif plasmid_inv_name in record.locus:
                        
                        found_seqs += 1
                        raw_seq = Seq(record.sequence)

                        # Locate START positions
                        start_1 = raw_seq.find(ss_1) + len(ss_1)
                        start_2 = raw_seq.find(ss_2) + len(ss_2)
                        
                        # Locate STOP positions
                        stop_1 = raw_seq.find(light_end) + len_light_end
                        stop_2 = raw_seq.find(heavy_end) + len_heavy_end
                        
                        #Extract light_end and heavy_end chains
                        light_seq = Seq(raw_seq[start_1:stop_1]).translate(to_stop=True, cds=False)
                        heavy_seq = Seq(raw_seq[start_2:stop_2]).translate(to_stop=True, cds=False)

                        # Create Protein object and conduct analysis
                        #TODO what if no seq match???
                        if light_seq and heavy_seq:
                            full_ab = Protein()
                            full_ab.add_sequence(2 * light_seq + 2 * heavy_seq)
                            full_ab.aa_distr()
                            full_ab.prot_mass()
                            full_ab.abs_coeff()
                            full_ab.pI()
                        
                            # TODO Refactor write data
                            
                            prot_name = ws_proteins.cell(
                                column=pt_header['POI name'],
                                row=i
                                ).value
                            
                            if prot_name is None:
                                prot_name = plasmid_inv_name.partition('.')[2]
                                ws_proteins.cell(
                                    column=pt_header['POI name'],
                                    row=i
                                    ).value = prot_name

                            # Add sequence length
                            ws_proteins.cell(
                                column=pt_header['Length'],
                                row=i
                                ).value = full_ab.seq_len
                            # Add mass of POI
                            ws_proteins.cell(
                                column=pt_header['MW'],
                                row=i
                                ).value = round(full_ab.mass, 2)
                            # Add pI of POI
                            ws_proteins.cell(
                                column=pt_header['pI'],
                                row=i
                                ).value = round(full_ab.pI, 2)
                            # Add absorbance coefficients
                            ws_proteins.cell(
                                column=pt_header['A0.1% (Ox)'],
                                row=i
                                ).value = round(full_ab.a_ox, 3)
                            ws_proteins.cell(
                                column=pt_header['A0.1% (Red)'],
                                row=i
                                ).value = round(full_ab.a_red, 3)
    
    print("DONE\n")
              
    save_workbook(wb, mypath, file)
    wb.close()
    
    print(f'{found_seqs} matching sequences found in GeneBank files')

    task_end()


# 4) Stock transfer

def create_import_file():
    """  """
    mypath, file = get_path_file('xlsx')
    if file is None:
        return None
    
    task_start(file)
    
    template_file = "Templates\LG_stock_transfer.xlsx"
    pball_connection(template_file)
    
    plasmids, boxes = load_pl_transfer(token, mypath, file)
    transfer_wb = load_workbook(join(pb_all, template_file))
    ws = transfer_wb['Sheet1']
    
    # TODO Refactor - variable names
    
    for i, (stock_id, stock_data) in enumerate(plasmids.items(), 2):

        box_name = stock_data['box_name']
        ws.cell(row=i, column=1).value = stock_id
        ws.cell(row=i, column=2).value = stock_data['name']
        ws.cell(row=i, column=25).value = boxes[box_name]['name']
        ws.cell(row=i, column=26).value = boxes[box_name]['rows']
        ws.cell(row=i, column=27).value = boxes[box_name]['cols']
        ws.cell(row=i, column=28).value = stock_data['stock_pos']
        ws.cell(row=i, column=29).value = boxes[box_name]['hierarchy']
        ws.cell(row=i, column=30).value = 'Plasmid'
        ws.cell(row=i, column=31).value = stock_data['inventory_name']
        ws.cell(row=i, column=32).value = stock_data['sysid']
    
    filename = file.replace('.xlsx', '_LG_import.xlsx')
    transfer_wb.save(join(mypath, filename))
    transfer_wb.close()
    
    print(f'Import file saved as {filename}')
    
    while True:
        if_import = input('Open Labguru import page? (Y/N) ').upper()
        if if_import == "Y":
            open_import_lg()
            break
        elif if_import == "N":
            break
        else:
            print("---< Wrong input. Try again >---".center(80))
        
    task_end()

def open_import_lg():
    """
    Opens LG page for xlsx import/update.
    """
    print('\nMake sure you are logged to Labguru via webbrowser')
    system('pause')
    system('explorer "https://my.labguru.com/system/json_imports/new?class=System%3A%3AStorage%3A%3AStock"')
    print('\nTransfer import file to Labguru import page')

def load_pl_transfer(token, mypath, file):
    """ Creates dict with stock: data necessary for transfer """

    def find_pos(boxes_data, box_name):

        if boxes_data[box_name]['free_pos']:
            position = boxes_data[box_name]['free_pos'].pop(0)
        else:
            position = None
        return position
    
    # Checks if box data is loaded into boxes_data and loads it if not
    def update_box_data(boxes_data, box_name, box_id):
        free_pos = boxes_data[box_name].get('free_pos', -1)
        if free_pos == -1:
            box_update = get_box_data(token, box_id)
            boxes_data[box_name].update(box_update)

    # Add 
    def add_stock_position(boxes_data, boxes_str, boxes_id, position, box_name):
        
        box_name, box_id = switch_to_id(boxes_id, boxes_str, box_name)

        # TODO - new_box = add_box() if no box found
        # box_name = new_box['name']
        # box_id = new_box['id']
        
        # Convert box position to int
        if isinstance(position, str):
            position = pos_to_int(position)
        
        # Box and position given
        if box_name is not None and position is not None:
            update_box_data(boxes_data, box_name, box_id)
            if position in boxes_data[box_name]['free_pos']:
                boxes_data[box_name]['free_pos'].remove(position)
                return box_name, position
        
        # Only box given or given position unavailable
        if box_name is not None:
            update_box_data(boxes_data, box_name, box_id)
            position = find_pos(boxes_data, box_name)
            if position is not None:
                return box_name, position

        # Find box with free position         
        for box in boxes_str.keys():
            box_id = boxes_str[box]
            update_box_data(boxes_data, box, box_id)
            position = find_pos(boxes_data, box)

            if position is not None:
                return box, position

        while True:
            new_box_choice = input('No free box found for at least one stock. Add a new box? (Y/N)').upper()
            if new_box_choice == 'Y':
                new_box = add_box()
                new_box_name = new_box['name']
                new_box_id = new_box['id']
                
                boxes_data[new_box_name] = new_box_id
                update_box_data(boxes_data, new_box_name, new_box_id)
                position = find_pos(boxes_data, new_box)
                
                return new_box_name, position
            
            elif new_box_choice == 'N':
                break
            else:
                print('---< Wrong input. Try again >---'.center(80))
        
        return None, None
    
    def transfer_validation(row):
        '''
        Validate row for sample transfer
        '''
        transfer = row[pl_header['Transfer'] - 1]

        cond_id = row[pl_header['ID'] - 1] is not None
        cond_transfer_str = isinstance(transfer, str)
        cond_transfer_y = False
        if cond_transfer_str:
            cond_transfer_y = transfer.upper() == "Y"
            
        conditions = [
            cond_id,
            cond_transfer_str,
            cond_transfer_y,
            ]

        return all(conditions)
    
    def switch_to_id(boxes_id, boxes_str, box_name):
        '''
        Returns name and id of box
        '''
        if isinstance(box_name, int):
            try:
                box_name, box_id = boxes_id[box_name], box_name
            except KeyError:
                print(f'---< {box_name} not matching any box ID in Z23.C storage >---'.center(80))
                box_name, box_id = None, None
        elif isinstance(box_name, str):
            try:
                box_id = boxes_str[box_name]
            except KeyError:
                print(f'---< {box_name} not found in Z23.C storage >---'.center(80))
                box_name, box_id = None, None
        else:
            box_name, box_id = None, None
            
        return box_name, box_id
    
    
    wb = load_workbook(join(mypath, file))
    ws_plasmids = wb['Plasmids']
    pl_header = sheet_header(ws_plasmids[1])
    
    # Dicts changing box name>id and id>name
    boxes_str, boxes_id = scan_storage(token, '1796')  # Storage: Z23.C (ID: 1796)
    boxes_list = sorted(list(boxes_str.items()), key=lambda x: x[0])
    boxes_data = {box_name: {'id': box_id} for box_name, box_id in boxes_list}

    stocks = {}
    i = 1
    
    for row in ws_plasmids.iter_rows(
        min_row=2,
        max_row=101,
        min_col=1,
        values_only=True
        ):
        
        i += 1

        if transfer_validation(row):

            stock_id = row[pl_header['ID'] - 1]
            stock_name = row[pl_header['Stock name (Plasmid)'] - 1]
            plasmid_sysid = row[pl_header['SysID'] - 1]
            plasmid_name = row[pl_header['Plasmid inventory name'] - 1]
            box_name = row[pl_header['New Box'] - 1]
            position = row[pl_header['New Position'] - 1]

            box_name, position = add_stock_position(
                boxes_data,
                boxes_str,
                boxes_id,
                position,
                box_name
                )
            
            if box_name is None or position is None:
                print(f'---< No position found for stock {stock_name} >---'.center())
                continue
            
            else:
                stock_data = {
                    'name': stock_name,
                    'box_name': box_name,
                    'box_id': boxes_data[box_name]['id'],
                    'stock_pos': position,
                    'sysid': plasmid_sysid,
                    'inventory_name': plasmid_name,     
                    }

                stocks[stock_id] = stock_data
                
                ws_plasmids.cell(
                    row=i,
                    column=pl_header['New Box']
                    ).value = box_name
                ws_plasmids.cell(
                    row=i,
                    column=pl_header['New Position']
                    ).value = pos_to_str(position)

    print(f'{len(stocks)} stocks ready for transfer')

    save_workbook(wb, mypath, file)
    wb.close()
    return [stocks, boxes_data]


# 5) Add proteins and stocks

def add_pt_stocks():
    """ """
    
    mypath, file = get_path_file('xlsx')
    if file is None:
        return None
    
    wb = load_workbook(join(mypath, file))
    ws_proteins = wb['Proteins']
    pt_header= sheet_header(ws_proteins[1])
    
    added_massage = ' < New protein entry added'
    prot_count = 0
    stock_count = 0
    
    task_start(file)
    
    # TODO Refactor - write_data
    for i in range(2, 502, 5):
        added = False
        
        plasmid_sysid = ws_proteins.cell(
            row=i,
            column=pt_header['Plasmid SysID']
            ).value
        plasmid_inv_name = ws_proteins.cell(
            row=i,
            column=pt_header['Plasmid inventory name']
            ).value
        
        if plasmid_sysid and plasmid_inv_name:
            prot_sysid = ws_proteins.cell(
                row=i,
                column=pt_header['POI SysID']
                ).value
            prot_id = ws_proteins.cell(
                row=i,
                column=pt_header['POI ID']
                ).value
            prot_name = ws_proteins.cell(
                row=i,
                column=pt_header['POI name']
                ).value
            concentration = ws_proteins.cell(
                row=i,
                column=pt_header['Concentration']
                ).value
            
            # Create protein name from plasmid name if none given
            if not prot_name:
                prot_name =  plasmid_inv_name.partition('.')[2]
                ws_proteins.cell(
                    row=i,
                    column=pt_header['POI name']
                    ).value = prot_name
                print(prot_name)
            
            # Create protein if protein name given but no protein ID
            if not prot_id:
                tag = str(ws_proteins.cell(
                    row=i,
                    column=pt_header['Tag']
                    ).value)
                mw = str(ws_proteins.cell(
                    row=i,
                    column=pt_header['MW']
                    ).value)
                purification_method = ws_proteins.cell(
                    row=i,
                    column=pt_header['Purification method']
                    ).value
                abs_ox = ws_proteins.cell(
                    row=i,
                    column=pt_header['A0.1% (Ox)']
                    ).value
                abs_red = ws_proteins.cell(
                    row=i,
                    column=pt_header['A0.1% (Red)']
                    ).value
                # TODO what if only one abs given
                if abs_ox and abs_red:
                    abs_coeff = f'{abs_ox:.3f} (Ox) / {abs_red:.3f} (Red)'
                else:
                    abs_coeff = ""

                prot_data = {
                    "name": prot_name,
                    # "description": "",
                    # "owner_id": "31",
                    # "alternative_name": "",
                    # "gene": "",
                    # "species": "",
                    # "mutations": "",
                    # "chemical_modifications": "",
                    "tag": tag,
                    "purification_method": purification_method,
                    "mw": mw,
                    "extinction_coefficient_280nm": abs_coeff,
                    # "storage_buffer": "",
                    # "storage_temperature": "",
                    # "sequence": ""
                }

                # Add protein and get its SysID and ID
                prot = add_protein(prot_data)

                # json_file = f'{prot_name}.json'
                # with open(json_file, 'w') as f:
                #     json.dump(prot, f, indent=4)
                
                if prot is None:
                    continue
                
                # TODO if status code ... else error
                added = True
                prot_count += 1
                
                # Write SysID and ID into the workbook
                try:
                    prot_id = prot['id']
                    prot_sysid = prot['auto_name']
                    prot_url = f'https://my.labguru.com/biocollections/proteins/{prot_id}'
                    ws_proteins.cell(
                        row=i,
                        column=pt_header['POI ID']
                        ).value = prot_id
                    ws_proteins.cell(
                        row=i,
                        column=pt_header['POI ID']
                        ).hyperlink = prot_url
                    ws_proteins.cell(
                        row=i,
                        column=pt_header['POI SysID']
                        ).value = prot_sysid
                    print(f'{prot_sysid:>8} (ID: {prot_id:>5}) {prot_name:<40}{added_massage if added else ""}')
                except Exception as e:
                    # print(e)
                    print(f'Error: {e}')
                            
            for j in range(
                pt_header['Stock 1'],
                pt_header['Stock 5']+1
                ):

                stock_id = ws_proteins.cell(
                    row=i,
                    column=j
                    ).value
                vol = ws_proteins.cell(
                    row=i+1,
                    column=j
                    ).value
                box_id = ws_proteins.cell(
                    row=i+2,
                    column=j
                    ).value
                description = ws_proteins.cell(
                    row=i,
                    column=pt_header['Description']
                    ).value
                
                if (vol is None) or stock_id or (box_id is None):
                    continue
                
                # TODO get storage_type based on storage_id (Box or Rack Cell)
                                
                stock_data= {
                    "name": str(prot_name),
                    "storage_id": int(box_id),
                    "storage_type": "System::Storage::Box",
                    "stockable_type": "Biocollections::Protein",
                    "stockable_id": int(prot_id),
                    "description": str(description),
                    # "barcode": "",
                    # "stored_by": "",
                    "concentration": str(concentration),
                    "concentration_prefix": "",
                    "concentration_unit_id": 9, # mg/mL
                    "concentration_exponent": "",
                    "concentration_remarks": "",
                    "volume": str(vol),
                    "volume_prefix": "",
                    "volume_unit_id": 8, # uL
                    "volume_exponent": "",
                    "volume_remarks": "",
                    # "weight": "",
                    # "weight_prefix": "",
                    # "weight_unit_id": 1,
                    # "weight_exponent": "25",
                    # "weight_remarks": "weight remarks"
                }
            
                # TODO add id, position when added
                # stock_data['sample_id'] = ws_proteins.cell(row=i, column=j).value
                
                stock = add_stock(stock_data) # TODO

                if stock is None:
                    continue
                
                stock_count += 1                
                stock_id = stock['id']
                box = stock['box']['name']
                pos = stock['position']
                stock_url = f'https://my.labguru.com/storage/stocks/{stock_id}'
                
                ws_proteins.cell(row=i, column=j).value = stock_id #stock_id
                ws_proteins.cell(row=i, column=j).hyperlink = stock_url
                ws_proteins.cell(row=i+3, column=j).value = box # box name
                ws_proteins.cell(row=i+4, column=j).value = pos # box position
                
                print(f'\tStock {j-23} (ID: {stock_id:>6}) - Box: {box}, Position: {pos}')

    save_workbook(wb, mypath, file)
    wb.close()

    print(f'Protein entries created: {prot_count}')
    print(f'Stocks added: {stock_count}')
    print()
    
    task_end()


# 6) Labels

def create_label_xlsx():
    
    mypath = getcwd()
    ext = 'xlsx'
    xlsx_list = scan_files(mypath, ext)
    file = choose_file(mypath, xlsx_list, ext)
    if file is None:
        return None
    
    wb = load_workbook(join(mypath, file), data_only=True)
    ws_proteins = wb['Proteins']
    pt_header = sheet_header(ws_proteins[1])
    
    wb_label = Workbook()
    ws_label = wb_label.active
    
    label_i = 1
    
    task_start(file)
    
    for i in range(2, 502, 5):
        for j in range(
                pt_header['Stock 1'],
                pt_header['Stock 5']+1):
            
            stock_name = ws_proteins.cell(
                row=i,
                column=pt_header['POI name']
                ).value
            stock_tag = ws_proteins.cell(
                row=i,
                column=pt_header['Tag']
                ).value
            stock_conc = ws_proteins.cell(
                row=i,
                column=pt_header['Concentration']
                ).value
            stock_id = ws_proteins.cell(column=j, row=i).value
            stock_vol = ws_proteins.cell(column=j, row=i+1).value
            
            if not stock_id:
                continue
            
            ws_label.cell(column=1, row=label_i).value = stock_name
            ws_label.cell(column=2, row=label_i).value = stock_id
            ws_label.cell(column=3, row=label_i).value = round(stock_conc, 3)
            ws_label.cell(column=4, row=label_i).value = round(stock_vol, 0)
            ws_label.cell(column=5, row=label_i).value = stock_tag
            
            label_i += 1
    wb.close()
    
    label_file = file.replace('.xlsx', '_labels.xlsx')
    wb_label.save(join(mypath, label_file))
    wb_label.close()
    print(f'File saved as {label_file}')
    
    task_end()


def load_viability():
    
    mypath, file = get_path_file('xlsx')
    mypath, file_csv = get_path_file('csv')
    
    print(file, file_csv)




if __name__ == '__main__':
    main()
    # load_viability()
